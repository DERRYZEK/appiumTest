import openpyxl


class HomePageData:
    test_HomePage_Data = [
        {"name": "Derrick Owusu Ansah", "password": 12345, "gender": "Male", "email": "derrick@gmail.com"},
        {"name": "Rahel Owusu Ansah", "password": 123456, "gender": "Female",
         "email": "rahel@gmail.com"}]

    @staticmethod
    def getData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\owusu\\PycharmProjects\\PythonDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
